import bdb

from app import *
from app.models.user import *
from app.models.users_songs import *
from config import Configuration

import psycopg2
import psycopg2.extras
import io
import xlwt
import os
from openpyxl import load_workbook


conn = psycopg2.connect(dbname=Configuration.DB_NAME,
                        user=Configuration.DB_USER,
                        password=Configuration.DB_PASS,
                        host=Configuration.DB_HOST)


@app.route('/import_excel', methods=['GET', 'POST'])
@login_required
def import_excel():
    if request.method == 'POST':

        loc = request.files['upload_excel']
        if loc != '':
            data = load_workbook(loc)
            sheet = data.active  # first page
            max_column = sheet.max_column

            l = 0
            for c in range(1, max_column, 2):
                if sheet[1][c].value is None:
                    l = c - 1
                    break

            for col in range(1, l + 1, 2):
                name = sheet[1][col].value
                year = sheet[2][col].value
                artist = sheet[3][col].value
                songs = []

                for i in range(sheet.max_row + 1):
                    if sheet[4 + i][col].value is None:
                        break
                    subsongs = []
                    subsongs.append(sheet[4 + i][col].value)
                    subsongs.append(sheet[4 + i][col + 1].value)
                    songs.append(subsongs)

                albums = db.session.query(Album).with_entities(Album.id, Album.name, Album.artist_id).all()

                als = []
                for a in albums:
                    ar = db.session.query(Artist).with_entities(Artist.id, Artist.firstname,
                                                                Artist.surname).filter(Artist.id == a.artist_id).all()
                    if None in list(*ar):
                        als.append([a.id, a.name, list(*ar)[0], list(*ar)[1]])
                    else:
                        als.append([a.id, a.name, list(*ar)[0], f'{list(*ar)[1]} {list(*ar)[2]}'])
                # print(als)

                for a in als:
                    ass = db.session.query(Album).filter(Album.artist_id==a[2]).with_entities(Album.name).all()
                    resy = []
                    for aaas in ass:
                        resy.append(*aaas)
                    if a[1] == name and a[3] == artist:
                        ids = db.session.query(Album).filter(Album.artist_id == a[2] and Album.name==name).with_entities(Album.id).first()
                        db.session.query(Song).filter(Song.album_id == int(*ids)).delete()
                        db.session.commit()

                        for song in songs:
                            s = Song(name=song[0], duration=song[1], album_id=int(*ids))
                            db.session.add(s)
                            db.session.commit()
                        break
                    elif a[3] == artist and name not in resy:
                        album = Album(name=name, release_year=year, artist_id=a[2])
                        db.session.add(album)
                        db.session.commit()
                        id = album.id

                        for song in songs:
                            s = Song(name=song[0], duration=song[1], album_id=id)
                            db.session.add(s)
                            db.session.commit()
                        break
        else:
            return render_template('sequrity/import.html', title='import_excel')

    return render_template('sequrity/import.html', title='import_excel')


@app.route('/export_excel', methods=['GET'])
@login_required
def export_excel():
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    cur.execute("SELECT albums.id, albums.name, release_year, firstname, surname " 
                "FROM albums inner join artists a on a.id = albums.artist_id")
    result = cur.fetchall()

    cur.execute("SELECT a.id, songs.name, (songs.duration) "
                "FROM songs inner join albums a on a.id = songs.album_id;")
    songs = cur.fetchall()

    output = io.BytesIO()
    workbook = xlwt.Workbook()
    sh = workbook.add_sheet('Альбом')
    sh.write(0, 0, 'Ім\'я')
    sh.write(1, 0, 'Рік релізу')
    sh.write(2, 0, 'Артист')
    sh.write(3, 0, 'Пісня')
    time_format = xlwt.XFStyle()
    time_format.num_format_str = '0h:mm:ss'

    idx = 0
    for row in result:
        sh.write(0, idx+1, row[1])
        sh.write(1, idx+1, row['release_year'])
        if row['surname'] is not None:
            sh.write(2, idx+1, row['firstname'] + row['surname'])
        else:
            sh.write(2, idx + 1, row['firstname'])

        idd = 0
        for s in songs:
            if s[0] == row[0]:
                sh.write(3+idd, idx+1, s[1])
                sh.write(3+idd, idx+2, s[2], time_format)
                idd += 1
        idx += 2

    workbook.save(output)
    output.seek(0)

    return Response(output,
                    mimetype="application/ms-excel",
                    headers={"Content-Disposition": "attachment;filename=export.xlsx;"})
